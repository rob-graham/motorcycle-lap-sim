"""Import the supplied AiM-derived Mallala workbook into canonical SI channels."""

from pathlib import Path
import math

import numpy as np

from .model import TelemetrySession

_STANDARD_GRAVITY_MPS2 = 9.80665

_REQUIRED_HEADERS = (
    "Time", "Distance on GPS Speed", "east", "north", "GPS Speed",
    "GPS LatAcc", "GPS LonAcc", "GPS Slope", "GPS Heading", "GPS Gyro",
    "GPS Latitude", "GPS Longitude", "RollRate", "PitchRate", "YawRate",
    "ECU RPM", "ECU GEAR", "ECU THROTTLE", "ECU TPS HAND", "Dist from Start",
)


def _number(value):
    if value is None or value == "":
        return math.nan
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"telemetry contains non-numeric value {value!r}") from exc


def _integer(value):
    if value is None or value == "":
        return 0
    numeric = _number(value)
    if not math.isfinite(numeric) or not numeric.is_integer():
        raise ValueError(f"telemetry integer channel contains invalid value {value!r}")
    return int(numeric)


def load_aim_workbook(path, *, sheet_name="Updated") -> TelemetrySession:
    """Load an AiM-derived Excel sheet with cached formulas and convert to SI.

    The supplied project workbook has a human-readable header row followed by a
    unit row. `data_only=True` deliberately consumes the workbook's cached
    east/north formula values; the importer does not recalculate or silently
    reinterpret the workbook's coordinate conversion.

    The raw `ECU GEAR` channel is preserved as floating point because the
    supplied AiM workbook contains interpolated fractional values during gear
    transitions. Stable integer gear classification belongs in later telemetry
    cleaning/validation rather than the file importer.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel telemetry import requires the optional 'telemetry' dependencies") from exc

    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"telemetry workbook has no sheet {sheet_name!r}")
    sheet = workbook[sheet_name]
    headers = tuple(cell.value for cell in sheet[1])
    missing = [name for name in _REQUIRED_HEADERS if name not in headers]
    if missing:
        raise ValueError(f"telemetry sheet is missing required headers: {', '.join(missing)}")
    column = {name: headers.index(name) for name in _REQUIRED_HEADERS}

    time_s = []
    distance_m = []
    east_m = []
    north_m = []
    speed_mps = []
    lateral_mps2 = []
    longitudinal_mps2 = []
    slope_rad = []
    heading_rad = []
    gps_gyro_radps = []
    latitude_deg = []
    longitude_deg = []
    roll_rate_radps = []
    pitch_rate_radps = []
    yaw_rate_radps = []
    engine_rpm = []
    gear_number = []
    ecu_throttle_rad = []
    hand_throttle_fraction = []
    distance_from_start_m = []
    markers = []
    lap_ids = []

    marker_index = 20 if len(headers) > 20 else None
    lap_index = 21 if len(headers) > 21 else None
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if all(value is None for value in row):
            continue
        value = lambda name: row[column[name]]
        time_s.append(_number(value("Time")))
        distance_m.append(_number(value("Distance on GPS Speed")))
        east_m.append(_number(value("east")))
        north_m.append(_number(value("north")))
        speed_mps.append(_number(value("GPS Speed")) / 3.6)
        lateral_mps2.append(_number(value("GPS LatAcc")) * _STANDARD_GRAVITY_MPS2)
        longitudinal_mps2.append(_number(value("GPS LonAcc")) * _STANDARD_GRAVITY_MPS2)
        slope_rad.append(math.radians(_number(value("GPS Slope"))))
        heading_rad.append(math.radians(_number(value("GPS Heading"))))
        gps_gyro_radps.append(math.radians(_number(value("GPS Gyro"))))
        latitude_deg.append(_number(value("GPS Latitude")))
        longitude_deg.append(_number(value("GPS Longitude")))
        roll_rate_radps.append(math.radians(_number(value("RollRate"))))
        pitch_rate_radps.append(math.radians(_number(value("PitchRate"))))
        yaw_rate_radps.append(math.radians(_number(value("YawRate"))))
        engine_rpm.append(_number(value("ECU RPM")))
        gear_number.append(_number(value("ECU GEAR")))
        ecu_throttle_rad.append(math.radians(_number(value("ECU THROTTLE"))))
        hand_throttle_fraction.append(_number(value("ECU TPS HAND")) / 100.0)
        distance_from_start_m.append(_number(value("Dist from Start")))
        marker = row[marker_index] if marker_index is not None and marker_index < len(row) else None
        markers.append(None if marker in (None, "") else str(marker))
        lap_value = row[lap_index] if lap_index is not None and lap_index < len(row) else None
        lap_ids.append(_integer(lap_value))

    array = lambda values: np.asarray(values, dtype=float)
    return TelemetrySession(
        array(time_s), array(distance_m), array(east_m), array(north_m), array(speed_mps),
        array(lateral_mps2), array(longitudinal_mps2), array(slope_rad), array(heading_rad),
        array(gps_gyro_radps), array(latitude_deg), array(longitude_deg), array(roll_rate_radps),
        array(pitch_rate_radps), array(yaw_rate_radps), array(engine_rpm), array(gear_number),
        array(ecu_throttle_rad), array(hand_throttle_fraction), array(distance_from_start_m),
        np.asarray(lap_ids, dtype=np.int64), tuple(markers), sheet_name)
