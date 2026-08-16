"""Import GPS quality indicators from an AiM-derived workbook sheet."""

from pathlib import Path
import math

import numpy as np

from .quality import GPSQualitySeries

_REQUIRED_HEADERS = ("Time", "GPS Nsat", "GPS PosAccuracy", "GPS SpdAccuracy")


def _number(value):
    if value is None or value == "":
        return math.nan
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"GPS quality data contains non-numeric value {value!r}") from exc


def load_aim_gps_quality(path, *, sheet_name="R6MallalaP4") -> GPSQualitySeries:
    """Load raw AiM GPS-quality channels and convert accuracy fields to SI.

    The supplied workbook stores ``GPS PosAccuracy`` in millimetres and
    ``GPS SpdAccuracy`` in kilometres per hour.  They are converted to metres
    and metres per second respectively.  No quality thresholds are imposed by
    the importer.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel telemetry import requires the optional 'telemetry' dependencies") from exc

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"telemetry workbook has no sheet {sheet_name!r}")
    sheet = workbook[sheet_name]
    headers = tuple(cell.value for cell in sheet[1])
    missing = [name for name in _REQUIRED_HEADERS if name not in headers]
    if missing:
        raise ValueError(f"GPS quality sheet is missing required headers: {', '.join(missing)}")
    column = {name: headers.index(name) for name in _REQUIRED_HEADERS}

    time_s = []
    satellites = []
    position_accuracy_m = []
    speed_accuracy_mps = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if all(value is None for value in row):
            continue
        time_s.append(_number(row[column["Time"]]))
        satellites.append(_number(row[column["GPS Nsat"]]))
        position_accuracy_m.append(_number(row[column["GPS PosAccuracy"]]) / 1000.0)
        speed_accuracy_mps.append(_number(row[column["GPS SpdAccuracy"]]) / 3.6)

    array = lambda values: np.asarray(values, dtype=float)
    return GPSQualitySeries(
        time_s=array(time_s),
        satellites=array(satellites),
        position_accuracy_m=array(position_accuracy_m),
        speed_accuracy_mps=array(speed_accuracy_mps),
        source_sheet=sheet_name,
    )
